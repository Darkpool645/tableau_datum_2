from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADER_ROW = 5          # la fila 5 trae los encabezados
FIRST_DATA_ROW = 6      # el detalle arranca en la 6
LAST_COL = 23           # hasta la columna W

# columnas que deben salir numéricas para Tableau
NUMERIC_COLS = ["Cantidad", "Precio", "Impuesto", "Total", "Costo", "Margen"]
PERCENT_COLS = ["% Utilidad", "% Margen"]
DATE_COLS = ["Fecha"]

# La fila 5 repite "Total" (venta y tiempos). Aquí se desambigua.
DUP_RENAME = {"Total": "Total Tiempo"}

FILE_RE = re.compile(r"reporte_ventas_(\d{4})-(\d{2})(?:_parcial-(\d{2}))?\.xlsx$", re.I)


# --------------------------------------------------------------------------- #
# Selección de archivos
# --------------------------------------------------------------------------- #
def pick_files(download_dir: Path) -> list[Path]:
    """
    Un archivo por mes. Si existe el mes completo y además parciales,
    gana el completo; entre parciales gana el que cubre más días.
    """
    best: dict[str, tuple[int, Path]] = {}
    for f in sorted(Path(download_dir).glob("reporte_ventas_*.xlsx")):
        if f.name.startswith("~$"):          # temporales de Excel
            continue
        m = FILE_RE.search(f.name)
        if not m:
            print(f"    ignoro (nombre inesperado): {f.name}")
            continue
        tag = f"{m.group(1)}-{m.group(2)}"
        rank = 99 if m.group(3) is None else int(m.group(3))  # completo = 99
        if tag not in best or rank > best[tag][0]:
            best[tag] = (rank, f)

    files = [p for _, p in (best[t] for t in sorted(best))]
    for t in sorted(best):
        print(f"    {t} -> {best[t][1].name}")
    return files


# --------------------------------------------------------------------------- #
# Lectura de un archivo
# --------------------------------------------------------------------------- #
def unique_header(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for n in names:
        seen[n] = seen.get(n, 0) + 1
        out.append(n if seen[n] == 1 else DUP_RENAME.get(n, f"{n}_{seen[n]}"))
    return out


def read_report(path: Path) -> pd.DataFrame:
    """
    Lee A5:W<fin> de un reporte de Datum.

    El detalle termina donde la columna 'Area' se queda vacía: después de eso
    el archivo trae el 'Total' y un resumen por Tipo que NO queremos.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(min_row=HEADER_ROW, max_col=LAST_COL, values_only=True)
    header = unique_header([str(c).strip() if c is not None else f"col_{i}"
                            for i, c in enumerate(next(rows), 1)])

    data = []
    for row in rows:
        if row[0] is None or str(row[0]).strip() == "":
            break                      # llegamos al Total / resumen
        data.append(row)
    wb.close()

    df = pd.DataFrame(data, columns=header)
    df.insert(0, "archivo_origen", path.name)
    m = FILE_RE.search(path.name)
    df.insert(1, "periodo", f"{m.group(1)}-{m.group(2)}" if m else "")
    return df


# --------------------------------------------------------------------------- #
# Limpieza
# --------------------------------------------------------------------------- #
def clean(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].map(lambda v: v.strip() if isinstance(v, str) else v)

    for col in NUMERIC_COLS:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    for col in PERCENT_COLS:
        if col in df:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.rstrip("%").replace({"": None, "None": None}),
                errors="coerce",
            ) / 100

    for col in DATE_COLS:
        if col in df:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    return df


# --------------------------------------------------------------------------- #
# Escritura
# --------------------------------------------------------------------------- #
def write_xlsx(df: pd.DataFrame, out: Path) -> Path:
    """
    Escribe en modo write_only: aguanta cientos de miles de filas sin
    tragarse la RAM. Encabezado en negritas y congelado.
    """
    if len(df) > 1_048_575:
        raise RuntimeError(
            f"{len(df):,} filas no caben en un .xlsx (límite de Excel: 1,048,575). "
            "Usa --out DATUM.csv o DATUM.parquet."
        )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("ventas")
    ws.freeze_panes = "A2"

    bold = Font(bold=True)
    head = []
    for name in df.columns:
        c = WriteOnlyCell(ws, value=name)
        c.font = bold
        head.append(c)
    ws.append(head)

    for col_i, name in enumerate(df.columns, 1):
        letter = get_column_letter(col_i)
        width = max(10, min(38, int(df[name].astype(str).str.len().head(2000).max() or 10) + 2))
        ws.column_dimensions[letter].width = max(width, len(str(name)) + 2)

    date_cols = {df.columns.get_loc(c) for c in DATE_COLS if c in df.columns}
    for row in df.itertuples(index=False, name=None):
        cells = []
        for i, v in enumerate(row):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                cells.append(None)
            elif isinstance(v, pd.Timestamp):
                cells.append(v.to_pydatetime())
            else:
                cells.append(v)
        ws.append(cells)

    wb.save(out)
    # el formato de fecha se aplica al final, sobre el archivo ya escrito
    if date_cols:
        wb2 = load_workbook(out)
        ws2 = wb2["ventas"]
        for i in date_cols:
            letter = get_column_letter(i + 1)
            for cell in ws2[letter][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        wb2.save(out)
    return out


# --------------------------------------------------------------------------- #
# API pública
# --------------------------------------------------------------------------- #
def consolidate(download_dir: Path = Path("downloads"),
                out: Path = Path("DATUM.xlsx"),
                drop_duplicates: bool = False) -> Path:
    files = pick_files(download_dir)
    if not files:
        raise RuntimeError(f"No hay .xlsx en {Path(download_dir).resolve()}")

    frames = []
    for f in files:
        df = read_report(f)
        print(f"    {f.name}: {len(df):,} filas")
        frames.append(df)

    full = pd.concat(frames, ignore_index=True)
    full = clean(full)

    # OJO: apagado por defecto. Datum repite legítimamente la misma línea
    # (mismo producto, misma nota, mismo segundo) cuando se piden varios.
    if drop_duplicates:
        keys = [c for c in ("#Nota", "Producto", "Fecha", "Usuario", "Cantidad")
                if c in full]
        before = len(full)
        full = full.drop_duplicates(subset=keys, keep="first")
        if before != len(full):
            print(f"    quité {before - len(full):,} duplicados por {keys}")

    if "Fecha" in full:
        full = full.sort_values("Fecha", kind="stable")

    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    suffix = out.suffix.lower()
    if suffix == ".csv":
        full.to_csv(out, index=False, encoding="utf-8-sig")
    elif suffix == ".parquet":
        full.to_parquet(out, index=False)
    elif suffix == ".xlsx":
        write_xlsx(full, out)
    else:
        raise ValueError(f"Formato no soportado: {suffix} (usa .csv/.parquet/.xlsx)")

    print(f"\n{len(full):,} filas x {len(full.columns)} columnas -> {out.resolve()}")
    if "Fecha" in full and full["Fecha"].notna().any():
        print(f"    rango: {full['Fecha'].min()} .. {full['Fecha'].max()}")
    return out


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Une los reportes de Datum en un solo archivo")
    ap.add_argument("--dir", type=Path, default=Path("downloads"))
    ap.add_argument("--out", type=Path, default=Path("DATUM.xlsx"))
    ap.add_argument("--dedup", action="store_true",
                    help="quita líneas idénticas (#Nota+Producto+Fecha+Usuario)")
    a = ap.parse_args()
    consolidate(a.dir, a.out, drop_duplicates=a.dedup)