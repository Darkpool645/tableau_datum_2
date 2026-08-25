"""
Último paso del pipeline: acomoda el consolidado en hojas, como DATUM_bueno.xlsx.

Entrada : el DataFrame (o el .xlsx/.csv/.parquet) que produce consolidator.py
Salida  : un libro con 4 hojas
            Hoja1          -> Alimentos y Bebidas (+ Tipo Conjunto / Tipo / Subtipo)
            Casa club      -> Area de negocio = Casa club      (+ Union)
            Campo de golf  -> Area de negocio = Campo de golf  (+ Union)
            Gastos         -> captura manual, se conserva del libro anterior

Notas de diseño
---------------
* En DATUM_bueno.xlsx las columnas Tipo y Subtipo son fórmulas de Excel. Aquí se
  calculan en Python con exactamente la misma lógica: openpyxl escribe fórmulas
  SIN valor en caché, y Tableau lee la caché, así que una fórmula recién escrita
  llegaría como NULL. El valor calculado es idéntico al que Excel cachea.
* Fecha se escribe como datetime real (no texto) y los % como número con formato
  de porcentaje, para que Tableau los tipe solo.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Reglas de negocio
# --------------------------------------------------------------------------- #
AYB = "Alimentos y Bebidas"
CASA_CLUB = "Casa club"
CAMPO_GOLF = "Campo de golf"

# Areas que no son punto de venta de A y B
AREA_OVERRIDE = {
    "Estética": CASA_CLUB,
    "Estetica": CASA_CLUB,
    "Performance Lab": CAMPO_GOLF,
}
# Proshop vende de los dos lados: se decide por el grupo de primer nivel
PROSHOP = "Proshop"
GRUPO_CAMPO_PREFIX = "CAMPO DE GOLF"

# Basura que Datum mete cuando un día no trae movimientos
AREAS_BASURA = {"", "no existen registros para mostrar."}

SIN_SUBSUBGRUPO_AYB = "-"       # Hoja1 usa "-"; las otras hojas lo dejan vacío

# Tipo: se evalúa en este orden, primer match gana (igual que el IF anidado)
TIPO_RULES = [
    ("bebidas", "Bebidas"),
    ("alimentos", "Alimentos"),
    ("tabaco", "Tabaco"),
    ("descuento", "Descuentos"),
    ("evento", "Eventos"),
    ("clases golf", "Clases golf"),
    ("modificador", "Modificadores"),
]
SUBTIPO_RULES = [
    ("con alcohol", "con alcohol"),
    ("sin alcohol", "sin alcohol"),
]
SUBTIPO_DEFAULT = "-"

HOJA_AYB = "Hoja1"
HOJA_GASTOS = "Gastos"

COLS_AYB = [
    "Area", "Grupo", "Subgrupo", "Sub subgrupo", "Producto",
    "Tipo Conjunto", "Tipo", "Subtipo", "Usuario",
    "Cantidad", "Precio", "Impuesto", "Total", "Costo", "Margen",
    "% Utilidad", "% Margen", "Fecha", "Area de negocio",
]
COLS_NEGOCIO = [
    "Area", "Grupo", "Subgrupo", "Sub subgrupo", "Producto",
    "Tipo", "Usuario",
    "Cantidad", "Precio", "Impuesto", "Total", "Costo", "Margen",
    "% Utilidad", "% Margen", "Fecha", "Area de negocio", "Union",
]

GASTOS_HEADER = ["Administración", "Gastos AyB", "Gastos campo de golf",
                 "Gastos casa club", "Fecha"]

# Formatos de celda
FMT = {
    "Cantidad": "#,##0",
    "Precio": "#,##0.00",
    "Impuesto": "#,##0.00",
    "Total": "#,##0.00",
    "Costo": "#,##0.00",
    "Margen": "#,##0.00",
    "% Utilidad": "0%",
    "% Margen": "0%",
    "Fecha": "yyyy-mm-dd hh:mm:ss",
}
FUENTE = "Arial"


# --------------------------------------------------------------------------- #
# Derivaciones
# --------------------------------------------------------------------------- #
def split_grupos(serie: pd.Series) -> pd.DataFrame:
    """
    'CAMPO DE GOLF > Rentas CG > Carritos' -> tres columnas.

    Ojo: NO se hace strip. DATUM_bueno.xlsx conserva los espacios alrededor del
    '>' ('CAMPO DE GOLF ', ' Rentas CG ', ' Carritos') y los filtros de Tableau
    ya están hechos sobre esos valores.
    """
    partes = serie.fillna("").astype(str).str.split(">", n=2, expand=True)
    for i in range(3):
        if i not in partes.columns:
            partes[i] = None
    partes = partes[[0, 1, 2]]
    partes.columns = ["Grupo", "Subgrupo", "Sub subgrupo"]
    return partes.replace({"": None})


def _match(texto: str | None, reglas, default=None):
    if not texto:
        return default
    bajo = str(texto).lower()
    for aguja, valor in reglas:
        if aguja in bajo:
            return valor
    return default


def derive_tipo(tipo_conjunto: pd.Series) -> pd.Series:
    """Equivale al IF(ISNUMBER(SEARCH(...))) anidado de la columna G."""
    return tipo_conjunto.map(lambda v: _match(v, TIPO_RULES, ""))


def derive_subtipo(tipo_conjunto: pd.Series) -> pd.Series:
    """Equivale al IF anidado de la columna H."""
    return tipo_conjunto.map(lambda v: _match(v, SUBTIPO_RULES, SUBTIPO_DEFAULT))


def derive_area_negocio(area: pd.Series, grupo: pd.Series) -> pd.Series:
    def resolver(a, g):
        a = (a or "").strip()
        if a in AREA_OVERRIDE:
            return AREA_OVERRIDE[a]
        if a == PROSHOP:
            g = (g or "").strip().upper()
            return CAMPO_GOLF if g.startswith(GRUPO_CAMPO_PREFIX) else CASA_CLUB
        return AYB

    return pd.Series([resolver(a, g) for a, g in zip(area, grupo)], index=area.index)


# --------------------------------------------------------------------------- #
# Armado de las hojas
# --------------------------------------------------------------------------- #
def build_sheets(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    df = df.copy()

    # 1. tirar la basura de días sin movimientos
    area_norm = df["Area"].fillna("").astype(str).str.strip()
    basura = area_norm.str.lower().isin(AREAS_BASURA)
    if basura.any():
        print(f"    descarto {basura.sum():,} filas sin Area válida")
    df = df[~basura].copy()
    df["Area"] = area_norm[~basura]

    # 2. Grupos -> Grupo / Subgrupo / Sub subgrupo
    col_grupos = "Grupos" if "Grupos" in df.columns else "Grupo"
    df[["Grupo", "Subgrupo", "Sub subgrupo"]] = split_grupos(df[col_grupos])

    # 3. columnas derivadas
    df["Usuario"] = df["Usuario"].fillna("").astype(str).str.strip().str.upper()
    df["Area de negocio"] = derive_area_negocio(df["Area"], df["Grupo"])
    df["Tipo Conjunto"] = df["Tipo"]
    df["Union"] = 1

    if "Fecha" in df:
        df = df.sort_values("Fecha", kind="stable")

    hojas: dict[str, pd.DataFrame] = {}

    # --- Hoja1: Alimentos y Bebidas ---
    ayb = df[df["Area de negocio"] == AYB].copy()
    ayb["Tipo"] = derive_tipo(ayb["Tipo Conjunto"])
    ayb["Subtipo"] = derive_subtipo(ayb["Tipo Conjunto"])
    ayb["Sub subgrupo"] = ayb["Sub subgrupo"].fillna(SIN_SUBSUBGRUPO_AYB)
    hojas[HOJA_AYB] = ayb[COLS_AYB]

    # --- Casa club / Campo de golf: conservan el Tipo original ---
    for hoja, negocio in ((CASA_CLUB, CASA_CLUB), (CAMPO_GOLF, CAMPO_GOLF)):
        hojas[hoja] = df[df["Area de negocio"] == negocio][COLS_NEGOCIO].copy()

    for nombre, hoja in hojas.items():
        print(f"    {nombre}: {len(hoja):,} filas")
    return hojas


# --------------------------------------------------------------------------- #
# Gastos (captura manual: se hereda, nunca se pisa)
# --------------------------------------------------------------------------- #
def read_gastos(fuente: Path | None) -> list[tuple]:
    if not fuente or not Path(fuente).exists():
        return []
    try:
        wb = load_workbook(fuente, read_only=True, data_only=True)
    except Exception as e:
        print(f"    no pude abrir {fuente} para heredar Gastos: {e}")
        return []
    if HOJA_GASTOS not in wb.sheetnames:
        wb.close()
        return []
    filas = [r for r in wb[HOJA_GASTOS].iter_rows(min_row=2, max_col=5, values_only=True)
             if any(v is not None for v in r)]
    wb.close()
    print(f"    Gastos: heredo {len(filas)} filas de {Path(fuente).name}")
    return filas


# --------------------------------------------------------------------------- #
# Escritura
# --------------------------------------------------------------------------- #
def _ancho(serie: pd.Series, nombre: str) -> int:
    crudo = serie.astype(str).str.len().head(2000).max()
    contenido = int(crudo) if pd.notna(crudo) else 10
    return max(10, min(38, contenido + 2), len(nombre) + 2)


def _escribe_hoja(wb: Workbook, nombre: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(nombre)
    ws.freeze_panes = "A2"

    negritas = Font(name=FUENTE, bold=True)
    normal = Font(name=FUENTE)

    cabecera = []
    for col in df.columns:
        c = WriteOnlyCell(ws, value=col)
        c.font = negritas
        cabecera.append(c)
    ws.append(cabecera)

    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = _ancho(df[col], col)

    formatos = [FMT.get(c) for c in df.columns]
    for fila in df.itertuples(index=False, name=None):
        celdas = []
        for valor, fmt in zip(fila, formatos):
            if valor is None or (isinstance(valor, float) and pd.isna(valor)):
                celdas.append(None)
                continue
            if isinstance(valor, pd.Timestamp):
                valor = valor.to_pydatetime()
            if fmt:
                c = WriteOnlyCell(ws, value=valor)
                c.number_format = fmt
                c.font = normal
                celdas.append(c)
            else:
                celdas.append(valor)
        ws.append(celdas)


def _escribe_gastos(wb: Workbook, filas: list[tuple]) -> None:
    ws = wb.create_sheet(HOJA_GASTOS)
    negritas = Font(name=FUENTE, bold=True)

    cabecera = []
    for nombre in GASTOS_HEADER:
        c = WriteOnlyCell(ws, value=nombre)
        c.font = negritas
        cabecera.append(c)
    ws.append(cabecera)

    for i, nombre in enumerate(GASTOS_HEADER, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(nombre) + 2)

    for fila in filas:
        celdas = []
        for j, valor in enumerate(fila):
            c = WriteOnlyCell(ws, value=valor)
            c.number_format = "yyyy-mm-dd" if j == 4 else "#,##0.00"
            c.font = Font(name=FUENTE)
            celdas.append(c)
        ws.append(celdas)


def write_workbook(hojas: dict[str, pd.DataFrame], out: Path,
                   gastos: list[tuple]) -> Path:
    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)

    for nombre, df in hojas.items():
        if len(df) > 1_048_575:
            raise RuntimeError(
                f"La hoja '{nombre}' trae {len(df):,} filas y no cabe en Excel "
                "(límite 1,048,575)."
            )

    wb = Workbook(write_only=True)
    for nombre in (HOJA_AYB, CASA_CLUB, CAMPO_GOLF):
        _escribe_hoja(wb, nombre, hojas[nombre])
    _escribe_gastos(wb, gastos)
    wb.save(out)
    return out


# --------------------------------------------------------------------------- #
# API pública
# --------------------------------------------------------------------------- #
def _load(src) -> pd.DataFrame:
    if isinstance(src, pd.DataFrame):
        return src
    src = Path(src)
    suffix = src.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(src, parse_dates=["Fecha"])
    if suffix == ".parquet":
        return pd.read_parquet(src)
    if suffix == ".xlsx":
        return pd.read_excel(src, sheet_name=0)
    raise ValueError(f"Formato no soportado: {suffix}")


def layout(src, out: Path = Path("DATUM_final.xlsx"),
           gastos_from: Path | None = None) -> Path:
    """
    Acomoda el consolidado en hojas y escribe el libro final.

    src         : DataFrame o ruta al consolidado (.xlsx/.csv/.parquet)
    out         : libro final
    gastos_from : de dónde heredar la hoja Gastos (por defecto, el propio `out`
                  si ya existe: así la captura manual sobrevive cada corrida)
    """
    df = _load(src)
    hojas = build_sheets(df)

    fuente_gastos = gastos_from if gastos_from is not None else out
    gastos = read_gastos(fuente_gastos)
    if not gastos:
        print("    Gastos: sin datos previos, dejo solo los encabezados "
              "(captúralos en Excel y la próxima corrida los conserva)")

    destino = write_workbook(hojas, Path(out), gastos)
    total = sum(len(h) for h in hojas.values())
    print(f"\n{total:,} filas repartidas en {len(hojas) + 1} hojas -> {destino.resolve()}")
    return destino


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Acomoda el consolidado en hojas")
    ap.add_argument("--src", type=Path, default=Path("DATUM.xlsx"))
    ap.add_argument("--out", type=Path, default=Path("DATUM_final.xlsx"))
    ap.add_argument("--gastos-from", type=Path, default=None,
                    help="libro del que se hereda la hoja Gastos (default: --out)")
    a = ap.parse_args()
    layout(a.src, a.out, a.gastos_from)